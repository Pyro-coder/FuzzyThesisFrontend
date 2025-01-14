import os
import time

import pandas as pd
import openpyxl
import shutil
from flask import Flask, render_template, request
from psychDiagnosis.psychopathy_main import generate_plots

app = Flask(__name__)

# Function to load data from the Excel file
def load_data_from_excel(file_path):
    spreadsheet = pd.ExcelFile(file_path)
    scores_data = spreadsheet.parse('Scores')
    words_data = spreadsheet.parse('Words')

    # Construct 'criteria'
    criteria = scores_data.rename(columns={
        'Factors': 'name',
        'Weights': 'default_importance',
        'Scoring': 'description'
    }).to_dict(orient='records')

    # Construct 'scoring_criteria'
    scoring_criteria = {}
    for row in words_data.itertuples(index=False, name=None):
        if pd.notnull(row[0]):  # Ensure factor name is not NaN
            scoring_criteria[row[0]] = [value for value in row[1:] if pd.notnull(value)]

    return criteria, scoring_criteria


# Save updated data to Excel
def save_updates_to_excel(file_path, updated_criteria):
    workbook = openpyxl.load_workbook(file_path)
    scores_sheet = workbook['Scores']

    for row in scores_sheet.iter_rows(min_row=2, max_row=scores_sheet.max_row):
        factor = row[0].value  # Assuming the first column is "Factors"
        for item in updated_criteria:
            if item['name'] == factor:
                row[1].value = item.get('selected_importance', row[1].value)  # Update weights
                row[2].value = item.get('selected_score', row[2].value)  # Update scoring

    workbook.save(file_path)


# Ensure a static folder for storing plots
PLOTS_DIR = os.path.join(app.root_path, 'static', 'plots')
os.makedirs(PLOTS_DIR, exist_ok=True)

@app.route('/', methods=['GET', 'POST'])
def index():
    # Reload fresh data from the Excel file
    EXCEL_FILE = 'psychDiagnosis/excel/PCLRWords.xlsx'
    criteria, scoring_criteria = load_data_from_excel(EXCEL_FILE)

    if request.method == 'POST':
        results = request.form.to_dict()

        # Extract updated scores and weights
        updated_scores = {key: value for key, value in results.items() if key.endswith('_score')}
        updated_importance = {key: value for key, value in results.items() if key.endswith('_importance')}

        # Update criteria
        updated_criteria = []
        for item in criteria:
            name = item['name']
            updated_item = item.copy()  # Create a fresh copy
            updated_item['selected_score'] = updated_scores.get(f"{name}_score", "N/A")
            updated_item['selected_importance'] = updated_importance.get(f"{name}_importance", "N/A")
            updated_criteria.append(updated_item)

        # Save updates to Excel
        save_updates_to_excel(EXCEL_FILE, updated_criteria)

        # Clear plots directory to avoid residual data
        if os.path.exists(PLOTS_DIR):
            shutil.rmtree(PLOTS_DIR)
        os.makedirs(PLOTS_DIR, exist_ok=True)

        # Generate plots
        plot_paths = generate_plots(PLOTS_DIR)
        plot_urls = [os.path.relpath(path, app.root_path) for path in plot_paths]

        return render_template('results.html', results=updated_criteria, plots=plot_urls)

    return render_template('index.html', criteria=criteria, scoring_criteria=scoring_criteria)


if __name__ == '__main__':
    app.run(debug=True)
