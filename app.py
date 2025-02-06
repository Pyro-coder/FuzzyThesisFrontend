import os
import time
import shutil
import sys
import requests
import pandas as pd
import openpyxl
from flask import Flask, abort, render_template, request
from psychDiagnosis.psychopathy_main import generate_plots
from dotenv import load_dotenv

# Load environment variables from .env if present
load_dotenv()

# Get the reCAPTCHA secret key from the environment
RECAPTCHA_SECRET_KEY = os.getenv("RECAPTCHA_SECRET_KEY")

if not RECAPTCHA_SECRET_KEY:
    raise ValueError("Missing reCAPTCHA secret key! Set the RECAPTCHA_SECRET_KEY environment variable.")

# Determine the directory where the executable or script is located
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Define the runs directory
RUNS_DIR = os.path.join(BASE_DIR, 'runs')
os.makedirs(RUNS_DIR, exist_ok=True)

def create_run_folder():
    run_timestamp = time.strftime("%Y%m%d-%H%M%S")
    run_folder = os.path.join(RUNS_DIR, run_timestamp)
    os.makedirs(run_folder, exist_ok=True)
    return run_folder

def copy_excel_to_run_folder(source_excel, run_folder):
    destination_excel = os.path.join(run_folder, 'PCLRWords.xlsx')
    shutil.copy(source_excel, destination_excel)
    return destination_excel

def save_plots_to_run_folder(plots_source_dir, run_folder):
    plots_destination_dir = os.path.join(run_folder, 'plots')
    os.makedirs(plots_destination_dir, exist_ok=True)
    for file_name in os.listdir(plots_source_dir):
        file_path = os.path.join(plots_source_dir, file_name)
        if os.path.isfile(file_path):
            shutil.copy(file_path, plots_destination_dir)
    return plots_destination_dir

app = Flask(__name__, template_folder='frontend/templates', static_folder='frontend/static')

# Function to load data from the Excel file
def load_data_from_excel(file_path):
    spreadsheet = pd.ExcelFile(file_path)
    scores_data = spreadsheet.parse('Scores')
    words_data = spreadsheet.parse('Words')

    criteria = scores_data.rename(columns={
        'Factors': 'name',
        'Weights': 'default_importance',
        'Scoring': 'description'
    }).to_dict(orient='records')

    scoring_criteria = {}
    for row in words_data.itertuples(index=False, name=None):
        if pd.notnull(row[0]):
            scoring_criteria[row[0]] = [value for value in row[1:] if pd.notnull(value)]

    return criteria, scoring_criteria

# Save updated data to Excel
def save_updates_to_excel(file_path, updated_criteria):
    workbook = openpyxl.load_workbook(file_path)
    scores_sheet = workbook['Scores']

    for row in scores_sheet.iter_rows(min_row=2, max_row=scores_sheet.max_row):
        factor = row[0].value
        for item in updated_criteria:
            if item['name'] == factor:
                row[1].value = item.get('selected_importance', row[1].value)
                row[2].value = item.get('selected_score', row[2].value)

    workbook.save(file_path)

PLOTS_DIR = os.path.join(app.root_path, 'frontend', 'static', 'plots')
os.makedirs(PLOTS_DIR, exist_ok=True)

@app.route('/', methods=['GET', 'POST'])
def index():
    EXCEL_FILE = os.path.join(os.path.dirname(__file__), 'psychDiagnosis', 'excel', 'PCLRWords.xlsx')
    EXCEL_FILE = os.path.abspath(EXCEL_FILE)
    criteria, scoring_criteria = load_data_from_excel(EXCEL_FILE)

    if request.method == 'POST':
        recaptcha_response = request.form.get('g-recaptcha-response')

        # Verify reCAPTCHA with Google
        recaptcha_verify_url = "https://www.google.com/recaptcha/api/siteverify"
        recaptcha_payload = {
            'secret': RECAPTCHA_SECRET_KEY,
            'response': recaptcha_response
        }
        recaptcha_result = requests.post(recaptcha_verify_url, data=recaptcha_payload).json()

        if not recaptcha_result.get('success'):
            abort(400, description="reCAPTCHA verification failed. Please try again.")

        results = request.form.to_dict()

        # Extract scores and importance values
        updated_scores = {key: value for key, value in results.items() if key.endswith('_score') and value.strip()}
        updated_importance = {key: value for key, value in results.items() if key.endswith('_importance') and value.strip()}

        if not updated_scores and not updated_importance:
            abort(400, description="Invalid submission: No meaningful data provided.")

        run_folder = create_run_folder()
        copied_excel_path = copy_excel_to_run_folder(EXCEL_FILE, run_folder)

        updated_criteria = []
        for item in criteria:
            name = item['name']
            updated_item = item.copy()
            updated_item['selected_score'] = updated_scores.get(f"{name}_score", "N/A")
            updated_item['selected_importance'] = updated_importance.get(f"{name}_importance", "N/A")
            updated_criteria.append(updated_item)

        save_updates_to_excel(copied_excel_path, updated_criteria)

        if os.path.exists(PLOTS_DIR):
            shutil.rmtree(PLOTS_DIR)
        os.makedirs(PLOTS_DIR, exist_ok=True)

        plot_paths = generate_plots(PLOTS_DIR, copied_excel_path)
        save_plots_to_run_folder(PLOTS_DIR, run_folder)

        plot_urls = [os.path.join('static', 'plots', os.path.basename(path)) for path in plot_paths]

        return render_template('results.html', results=updated_criteria, plots=plot_urls)

    return render_template('index.html', criteria=criteria, scoring_criteria=scoring_criteria)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000)
